import pyperclip 
import openpyxl
import pyautogui

workbook = openpyxl.load_workbook('F:/Repo_GitHub/Codes_in_Py/auto preenchimento/produtos_ficticios.xlsx')
print(workbook.sheetnames)

sheet_produto = workbook['Produtos']

for linha in sheet_produto.iter_rows(min_row=2):
    nome_produto = linha[0].value
    pyperclip.copy(nome_produto)
    pyautogui.click(x=375, y=335,duration=2)
    pyautogui.hotkey('ctrl', 'v')
    
    descricao_produto = linha[1].value
    pyperclip.copy(descricao_produto)
    pyautogui.click(x=366, y=434,duration=2)
    pyautogui.hotkey('ctrl', 'v')
    
    categoria_produto = linha[2].value
    pyperclip.copy(categoria_produto)
    pyautogui.click(x=365, y=560,duration=2)
    pyautogui.hotkey('ctrl', 'v')
    
    codigo_produto = linha[3].value
    pyperclip.copy(codigo_produto)
    pyautogui.click(x=383, y=651,duration=2)
    pyautogui.hotkey('ctrl', 'v')
    
    peso = linha[4].value
    pyperclip.copy(peso)
    pyautogui.click(x=382, y=725,duration=2)
    pyautogui.hotkey('ctrl', 'v')
    
    dimesoes = linha[5].value
    pyperclip.copy(dimesoes)
    pyautogui.click(x=416, y=820,duration=2)
    pyautogui.hotkey('ctrl', 'v')
    
    pyautogui.click(x=351, y=869,duration=2)
    
    preco = linha[6].value
    quantidade = linha[7].value
    data_validade = linha[8].value
    cor = linha[9].value
    tamanho = linha[10].value
    material = linha[11].value
    fabricante = linha[12].value
    